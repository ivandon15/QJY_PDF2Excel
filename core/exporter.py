"""
core/exporter.py — 导出带备注列的 Excel
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

_COLS = [
    ("url",          "URL"),
    ("title",        "标题"),
    ("views",        "播放量"),
    ("danmaku",      "弹幕量"),
    ("date",         "发布时间"),
    ("has_playlist", "是否有合集"),
    ("likes",        "点赞量"),
    ("coins",        "投币量"),
    ("favorites",    "收藏量"),
    ("shares",       "转发量"),
    ("comments",     "评论数"),
    ("description",  "视频描述"),
    ("tags",         "视频标签"),
    ("notes",        "备注"),
]

_HEADER_FILL = PatternFill("solid", fgColor="2E4057")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
_URL_FONT = Font(color="1155CC", underline="single", name="Arial")
_BODY_FONT = Font(name="Arial")
_ALIGN_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=False)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def export_excel(rows: list[dict], output_path: str):
    """将聚合后的行列表导出为 Excel 文件。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "提取结果"

    # 表头
    for col_num, (_, label) in enumerate(_COLS, 1):
        cell = ws.cell(row=1, column=col_num, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _ALIGN_CENTER

    # 冻结首行
    ws.freeze_panes = "A2"

    # 数据行
    for row_num, row in enumerate(rows, 2):
        for col_num, (key, _) in enumerate(_COLS, 1):
            val = row.get(key, "")
            cell = ws.cell(row=row_num, column=col_num, value=val)
            if key == "url" and val:
                cell.hyperlink = val
                cell.font = _URL_FONT
            else:
                cell.font = _BODY_FONT
            cell.alignment = _ALIGN_LEFT

    # 列宽自适应
    col_max_lens = {}
    for row_num in range(1, ws.max_row + 1):
        for col_num in range(1, len(_COLS) + 1):
            v = ws.cell(row=row_num, column=col_num).value or ""
            cur = col_max_lens.get(col_num, 0)
            # 中文字符按2倍计算宽度
            length = sum(2 if ord(c) > 127 else 1 for c in str(v))
            col_max_lens[col_num] = max(cur, length)

    FIXED_WIDTHS = {
        1: 45,   # URL
        5: 16,   # 发布时间
        6: 12,   # 是否有合集
        12: 60,  # 视频描述
        13: 40,  # 视频标签
        14: 50,  # 备注
    }
    for col_num in range(1, len(_COLS) + 1):
        if col_num in FIXED_WIDTHS:
            ws.column_dimensions[get_column_letter(col_num)].width = FIXED_WIDTHS[col_num]
        else:
            w = min(col_max_lens.get(col_num, 8) + 4, 30)
            ws.column_dimensions[get_column_letter(col_num)].width = max(w, 8)

    wb.save(output_path)
