from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def export_to_excel(rows: list[dict], output_path: str):
    if not rows:
        return

    # Collect all keys in order, page first
    keys = ["page"]
    for row in rows:
        for k in row:
            if k != "page" and k not in keys:
                keys.append(k)

    wb = Workbook()
    ws = wb.active
    ws.title = "提取结果"

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for col, k in enumerate(keys, 1):
        cell = ws.cell(row=1, column=col, value=k)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        for col, k in enumerate(keys, 1):
            ws.cell(row=row_idx, column=col, value=row.get(k, ""))

    for col in range(1, len(keys) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 60)

    wb.save(output_path)
